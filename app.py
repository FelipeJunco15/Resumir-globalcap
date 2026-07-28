from datetime import datetime
import os
import pandas as pd
from flask import Flask, request
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# =========================================================================
# CONFIGURACIÓN GLOBAL DEL SERVIDOR Y CARPETAS
# =========================================================================
app = Flask(__name__)
CORS(app) 

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "salidas_prueba"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# =========================================================================
# 1. ENDPOINT: WILLIAM ENERCARE
# =========================================================================
@app.route('/upload/william', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "No hay archivo", 400

    file = request.files['file']
    if file.filename == '':
        return "Nombre de archivo vacío", 400

    path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(path)

    try:
        resultados_finales = ""
        dfs_exportacion = []

        with pd.ExcelFile(path, engine="openpyxl") as excel:
            for hoja in excel.sheet_names:
                try:
                    sheet_state = excel.book[hoja].sheet_state
                    if sheet_state == "hidden" or sheet_state == "veryHidden":
                        continue

                    df = pd.read_excel(excel, sheet_name=hoja, header=None)
                    if df.empty:
                        continue

                    fila_fechas_idx = 1
                    columna_nombres_idx = 0
                    
                    for col_idx in range(len(df.columns)):
                        val = str(df.iloc[fila_fechas_idx, col_idx]).strip()
                        if val == 'Week' or val == 'Week#':
                            columna_nombres_idx = col_idx
                            break

                    columna_nombres = df.iloc[:, columna_nombres_idx].astype(str).str.strip().str.lower()

                    filas_mapeadas = {
                        "Forecast": None,
                        "Required": None,
                        "Projected": None
                    }

                    for idx, text_celda in columna_nombres.items():
                        if text_celda in ["base fte required", "base fte required budget - delivered", "base fte required - 100%"]:
                            filas_mapeadas["Forecast"] = df.iloc[idx].copy()
                        elif text_celda in ["hc required", "required hc budget - with shrink", "required fte", "required hc"]:
                            filas_mapeadas["Required"] = df.iloc[idx].copy()
                        elif text_celda == "projected hc":
                            filas_mapeadas["Projected"] = df.iloc[idx].copy()

                    if all(v is None for v in filas_mapeadas.values()):
                        continue

                    filas_finales = [df.iloc[fila_fechas_idx].copy()]
                    for nombre_estandar, fila_datos in filas_mapeadas.items():
                        if fila_datos is not None:
                            fila_datos.iloc[columna_nombres_idx] = nombre_estandar
                            filas_finales.append(fila_datos)

                    df_resultado_hoja = pd.DataFrame(filas_finales)
                    df_resultado_hoja.columns = df_resultado_hoja.iloc[0]
                    df_resultado_hoja = df_resultado_hoja[1:]
                    
                    col_guia = df_resultado_hoja.columns[columna_nombres_idx]
                    df_resultado_hoja.rename(columns={col_guia: 'Fecha'}, inplace=True)

                    columnas_finales_2026 = ['Fecha']
                    for col in df_resultado_hoja.columns:
                        if col == 'Fecha':
                            continue
                        col_str = str(col).strip()
                        if col_str in ['None', 'nan', ''] or col_str.startswith('Unnamed:'):
                            continue
                        try:
                            fecha_parseada = pd.to_datetime(col, errors='coerce')
                            if pd.notna(fecha_parseada) and fecha_parseada.year == 2026:
                                columnas_finales_2026.append(col)
                        except:
                            if '2026' in col_str or '-26' in col_str:
                                columnas_finales_2026.append(col)

                    df_resultado_hoja = df_resultado_hoja[columnas_finales_2026]
                    df_resultado_hoja.dropna(how='all', axis=1, inplace=True)

                    if len(df_resultado_hoja.columns) <= 1:
                        continue

                    nuevos_nombres_cabecera = {}
                    for col in df_resultado_hoja.columns:
                        if col != 'Fecha':
                            try:
                                f_parsed = pd.to_datetime(col, errors='coerce')
                                if pd.notna(f_parsed):
                                    nuevos_nombres_cabecera[col] = f_parsed.strftime('%d-%b-%y')
                            except:
                                pass
                    
                    df_resultado_hoja.rename(columns=nuevos_nombres_cabecera, inplace=True)

                    for col in df_resultado_hoja.columns:
                        if col == 'Fecha':
                            continue
                        def forzar_entero_exacto(valor):
                            if pd.isna(valor) or str(valor).strip() == "":
                                return ""
                            try:
                                return int(round(float(valor)))
                            except:
                                return valor
                        df_resultado_hoja[col] = df_resultado_hoja[col].apply(forzar_entero_exacto)

                    df_temp_export = df_resultado_hoja.copy()
                    df_temp_export.insert(0, "Pestaña", hoja)
                    dfs_exportacion.append(df_temp_export)
                    
                    resultados_finales += f"<h2>Pestaña: {hoja}</h2>"
                    resultados_finales += df_resultado_hoja.to_html(index=False, border=1, na_rep="")
                    resultados_finales += "<br><hr>"

                except Exception as e_hoja:
                    print(f" Error en pestaña '{hoja}': {e_hoja}")
                    continue

        if dfs_exportacion:
            df_excel_final = pd.concat(dfs_exportacion, ignore_index=True)
            df_excel_final.rename(columns={"Pestaña": "Program H2R"}, inplace=True)
            df_excel_final.insert(1, "Client", "Enercare")
            
            nombre_base = os.path.splitext(file.filename)[0]
            ruta_excel_salida = os.path.join(OUTPUT_FOLDER, f"{nombre_base}_Resumen_William_enercare.xlsx")
            
            with pd.ExcelWriter(ruta_excel_salida, engine="openpyxl") as writer:
                df_excel_final.to_excel(writer, sheet_name="Program H2R", index=False)

        if not resultados_finales:
            return "<h3>No se encontraron datos del año 2026 para las métricas solicitadas en William.</h3>"

        return f"<div>{resultados_finales}</div>"

    except Exception as e:
        return f"Error al leer datos de William: {str(e)}", 500


# =========================================================================
# 2. ENDPOINT: SURESH - FORMATO 1 (Métrica Vertical / Fechas Horizontales)
# =========================================================================
@app.route('/upload/suresh', methods=['POST'])
def procesar_suresh_individual():
    try:
        file = request.files.get('file')
        if not file or file.filename == '':
            return "<b style='color:red;'>Error: Archivo no recibido.</b>", 400
            
        nombre_base = os.path.splitext(file.filename)[0]
        ruta_salida_excel = os.path.join(OUTPUT_FOLDER, f"{nombre_base}_Resumen_Suresh_F1.xlsx")
        
        ruta_guardado = os.path.join(UPLOAD_FOLDER, f"temporal_{file.filename}")
        file.save(ruta_guardado)
        
        wb_final = Workbook()
        ws_final = wb_final.active
        ws_final.title = "Resumen Continuous 2026"
        
        columnas_finales = ['H2R Department', 'Client', 'Country', 'Fecha', 'Budget', 'Forecast', 'Required', 'Projected']
        ws_final.append(columnas_finales)
        
        for num_c in range(1, len(columnas_finales) + 1):
            ws_final.cell(row=1, column=num_c).font = Font(name="Calibri", size=11, bold=True)
            
        wb_origen = load_workbook(ruta_guardado, data_only=True)
        contador_filas_guardadas = 0
        
        for nombre_hoja in wb_origen.sheetnames:
            if any(p in nombre_hoja.lower() for p in ["geo", "reference", "sheet", "hoja"]) or nombre_hoja.startswith("Sheet"):
                continue
            ws = wb_origen[nombre_hoja]
            if ws.sheet_state in ["hidden", "veryHidden"]:
                continue
                
            filas_celdas = [list(r) for r in ws.iter_rows(values_only=True)]
            if len(filas_celdas) < 2: continue
                
            fila_1 = filas_celdas[0]
            columnas_2026 = {}
            for idx_col, valor_celda in enumerate(fila_1):
                if valor_celda is None: continue
                if isinstance(valor_celda, datetime) and valor_celda.year == 2026:
                    columnas_2026[idx_col] = valor_celda.strftime('%d-%b-%y')
                else:
                    texto_celda = str(valor_celda).strip()
                    if '2026' in texto_celda or '-26' in texto_celda or '/26' in texto_celda:
                        try:
                            fecha_convertida = pd.to_datetime(texto_celda)
                            if fecha_convertida.year == 2026:
                                columnas_2026[idx_col] = fecha_convertida.strftime('%d-%b-%y')
                        except:
                            columnas_2026[idx_col] = texto_celda
            
            if not columnas_2026: continue
                
            h2r_memoria = ""
            cliente_memoria = ""
            pais_memoria = str(nombre_hoja).strip()
            matriz_datos = {}
            
            for num_fila in range(2, len(filas_celdas) + 1):
                if num_fila in ws.row_dimensions and ws.row_dimensions[num_fila].hidden: continue
                fila_actual = filas_celdas[num_fila - 1]
                if not fila_actual: continue
                while len(fila_actual) < max(list(columnas_2026.keys()) + [5]): fila_actual.append(None)
                    
                c_h2r = str(fila_actual[0]).strip() if fila_actual[0] is not None else ""
                c_cli = str(fila_actual[1]).strip() if fila_actual[1] is not None else ""
                c_met = str(fila_actual[3]).strip() if fila_actual[3] is not None else ""
                
                if c_h2r != "" and c_h2r.lower() != "nan" and "program" not in c_h2r.lower(): h2r_memoria = c_h2r
                if c_cli != "" and c_cli.lower() != "nan" and "client" not in c_cli.lower(): cliente_memoria = c_cli
                    
                if h2r_memoria == "" or h2r_memoria.lower() in ["nan", "desconocido", "none"]: continue
                    
                metrica_limpia = c_met.strip().lower()
                if 'vs' in metrica_limpia or 'actual' in metrica_limpia or metrica_limpia == "": continue
                    
                nombre_metrica_final = None
                if 'bud' in metrica_limpia: nombre_metrica_final = 'Budget'
                elif 'fore' in metrica_limpia or 'cast' in metrica_limpia: nombre_metrica_final = 'Forecast'
                elif 'req' in metrica_limpia or 'fte' in metrica_limpia or 'requer' in metrica_limpia: nombre_metrica_final = 'Required'
                elif 'proj' in metrica_limpia or 'plan' in metrica_limpia: nombre_metrica_final = 'Projected'
                else: continue
                    
                if nombre_metrica_final:
                    for idx_col_fecha, texto_fecha in columnas_2026.items():
                        if idx_col_fecha >= len(fila_actual): continue
                        valor_celda_numero = fila_actual[idx_col_fecha]
                        val_str = str(valor_celda_numero).strip().lower()
                        
                        if val_str in ['', '-', 'nan', 'closed', '#value!', '#¡valor!', '#¡div/0!', 'none', 'null']:
                            valor_numerico = 0
                        else:
                            try: valor_numerico = int(round(float(valor_celda_numero)))
                            except: valor_numerico = 0
                                
                        llave = (h2r_memoria, cliente_memoria, texto_fecha)
                        if llave not in matriz_datos:
                            matriz_datos[llave] = {'Budget': 0, 'Forecast': 0, 'Required': 0, 'Projected': 0, 'tiene_datos_reales': False}
                        matriz_datos[llave][nombre_metrica_final] = valor_numerico
                        matriz_datos[llave]['tiene_datos_reales'] = True
            
            for (campana, cliente, fecha_texto), metricas in matriz_datos.items():
                if not metricas.get('tiene_datos_reales', False): continue
                ws_final.append([campana, cliente, pais_memoria, fecha_texto, metricas['Budget'], metricas['Forecast'], metricas['Required'], metricas['Projected']])
                contador_filas_guardadas += 1
                
        wb_origen.close()
        try: os.remove(ruta_guardado)
        except: pass

        if contador_filas_guardadas > 0:
            wb_final.save(ruta_salida_excel)
            return f"<h2 style='color:#27ae60; margin:0;'>¡Resumen Exitoso Formato 1!</h2><p>Guardado en salidas_prueba.</p>"
        return "<h3 style='color:#e67e22;'>Aviso: No se encontraron registros procesables para 2026.</h3>", 200
    except Exception as e:
        return f"<b style='color:red;'>Error Crítico F1: {str(e)}</b>", 200


# =========================================================================
# 3. ENDPOINT: SURESH - FORMATO 2 (Métrica Horizontal / Fechas Horizontales)
# =========================================================================
@app.route('/upload/suresh2', methods=['POST'])
def procesar_suresh_2_unico():
    try:
        file = request.files.get('file')
        if not file or file.filename == '':
            return "<b style='color:red;'>Error: No seleccionaste ningún archivo.</b>", 400
            
        nombre_base = os.path.splitext(file.filename)[0]
        ruta_salida_excel = os.path.join(OUTPUT_FOLDER, f"{nombre_base}_Resumen_Suresh_F2.xlsx")
        
        ruta_guardado = os.path.join(UPLOAD_FOLDER, f"temporal_{file.filename}")
        file.save(ruta_guardado)
        
        wb_final = Workbook()
        ws_final = wb_final.active
        ws_final.title = "Resumen Continuo 2026"
        
        columnas_finales = ['Program H2R', 'Client Name', 'Country', 'Fecha', 'Budget', 'Forecast', 'Required', 'Projected']
        ws_final.append(columnas_finales)
        
        for num_c in range(1, len(columnas_finales) + 1):
            ws_final.cell(row=1, column=num_c).font = Font(name="Calibri", size=11, bold=True)
            
        wb_origen = load_workbook(ruta_guardado, data_only=True)
        contador_filas_guardadas = 0
        
        for nombre_hoja in wb_origen.sheetnames:
            if any(p in nombre_hoja.lower() for p in ["geo", "reference", "sheet", "hoja"]) or nombre_hoja.startswith("Sheet"):
                continue
            ws = wb_origen[nombre_hoja]
            if ws.sheet_state in ["hidden", "veryHidden"]: continue
                
            filas_celdas = [list(r) for r in ws.iter_rows(values_only=True)]
            if len(filas_celdas) < 3: continue
                
            fila_cabecera = filas_celdas[2]
            columnas_2026 = {}
            idx_col_cliente = 1
            idx_col_pais = 2
            
            for idx_col, valor_celda in enumerate(fila_cabecera):
                if valor_celda is None: continue
                texto_celda = str(valor_celda).strip()
                texto_celda_min = texto_celda.lower()
                
                if "name(campaign)" in texto_celda_min:
                    idx_col_cliente = idx_col
                    continue
                elif "country" in texto_celda_min:
                    idx_col_pais = idx_col
                    continue
                
                if '2026' in texto_celda or '-26' in texto_celda or '/26' in texto_celda:
                    try:
                        fecha_dt = pd.to_datetime(texto_celda)
                        columnas_2026[idx_col] = fecha_dt.strftime('%d-%b-%y')
                    except:
                        columnas_2026[idx_col] = texto_celda
                elif isinstance(valor_celda, datetime) and valor_celda.year == 2026:
                    columnas_2026[idx_col] = valor_celda.strftime('%d-%b-%y')
            
            if not columnas_2026: continue
                
            h2r_memoria = ""
            cliente_memoria = ""
            pais_memoria = ""
            matriz_datos = {}
            
            for num_fila in range(4, len(filas_celdas) + 1):
                if num_fila in ws.row_dimensions and ws.row_dimensions[num_fila].hidden: continue
                fila_actual = filas_celdas[num_fila - 1]
                if not fila_actual: continue
                    
                while len(fila_actual) < max(list(columnas_2026.keys()) + [10]): fila_actual.append(None)
                    
                val_h2r = str(fila_actual[0]).strip() if fila_actual[0] is not None else ""
                val_cli = str(fila_actual[idx_col_cliente]).strip() if fila_actual[idx_col_cliente] is not None else ""
                val_pais = str(fila_actual[idx_col_pais]).strip() if fila_actual[idx_col_pais] is not None else ""
                
                if val_h2r != "" and "h2r department" not in val_h2r.lower(): h2r_memoria = val_h2r
                if val_cli != "" and "name(campaign)" not in val_cli.lower(): cliente_memoria = val_cli
                if val_pais != "" and "country" not in val_pais.lower(): pais_memoria = val_pais
                    
                nombre_metrica_final = None
                for c_idx in range(3, 6):
                    if c_idx >= len(fila_actual) or fila_actual[c_idx] is None: continue
                    txt_m = str(fila_actual[c_idx]).strip()
                    if txt_m == "Budget": nombre_metrica_final = 'Budget'; break
                    elif txt_m == "Forecast": nombre_metrica_final = 'Forecast'; break
                    elif txt_m == "Required": nombre_metrica_final = 'Required'; break
                    elif txt_m == "Projected": nombre_metrica_final = 'Projected'; break
                
                if not nombre_metrica_final: continue
                if cliente_memoria == "":
                    if h2r_memoria != "": cliente_memoria = h2r_memoria
                    else: continue
                if h2r_memoria == "": h2r_memoria = cliente_memoria
                
                for idx_col_fecha, texto_fecha in columnas_2026.items():
                    valor_celda_numero = fila_actual[idx_col_fecha]
                    val_str = str(valor_celda_numero).strip().lower()
                    
                    if val_str in ['closed', 'closd', 'clsoed']:
                        valor_numerico = "Closed"
                    elif val_str in ['', '-', 'nan', '#value!', '#¡valor!', '#¡div/0!', 'none', 'null']:
                        valor_numerico = 0
                    else:
                        try: valor_numerico = int(round(float(valor_celda_numero)))
                        except: valor_numerico = 0
                            
                    llave = (h2r_memoria, cliente_memoria, pais_memoria, texto_fecha)
                    if llave not in matriz_datos:
                        matriz_datos[llave] = {'Budget': 0, 'Forecast': 0, 'Required': 0, 'Projected': 0, 'tiene_datos': False}
                    matriz_datos[llave][nombre_metrica_final] = valor_numerico
                    matriz_datos[llave]['tiene_datos'] = True
            
            for (h2r, cli, pais, fecha_texto), metricas in matriz_datos.items():
                if not metricas['tiene_datos']: continue
                ws_final.append([h2r, cli, pais, fecha_texto, metricas['Budget'], metricas['Forecast'], metricas['Required'], metricas['Projected']])
                contador_filas_guardadas += 1
                
        wb_origen.close()
        try: os.remove(ruta_guardado)
        except: pass

        if contador_filas_guardadas > 0:
            wb_final.save(ruta_salida_excel)
            return f"<h2 style='color:#3498db; margin:0;'>¡Resumen Exitoso Formato 2!</h2><p>Guardado en salidas_prueba.</p>"
        return "<h3 style='color:#e67e22;'>No se encontraron columnas procesables.</h3>", 200
    except Exception as e:
        return f"<b style='color:red;'>Error Crítico F2: {str(e)}</b>", 200
    


























@app.route('/upload/miranda', methods=['POST'])
def upload_miranda():
    try:
        file = request.files.get('file')
        if not file or file.filename == '':
            return "No se seleccionó ningún archivo.", 400
            
        carpeta_uploads = "uploads"
        if not os.path.exists(carpeta_uploads):
            os.makedirs(carpeta_uploads, exist_ok=True)
            
        ruta_guardado = os.path.join(carpeta_uploads, file.filename)
        file.save(ruta_guardado)
        
        print(f"--- Procesando Reporte Miranda: {file.filename} ---")
        
        dataframes_totales = []
        
        # 1. CARGAR CON OPENPYXL PRIMERO PARA IDENTIFICAR FILAS FILTRADAS/ESCONDIDAS
        wb_control = load_workbook(ruta_guardado, data_only=True)
        
        with pd.ExcelFile(ruta_guardado, engine="openpyxl") as excel:
            for hoja in excel.sheet_names:
                sheet_state = excel.book[hoja].sheet_state
                if sheet_state == "hidden" or sheet_state == "veryHidden" or hoja == "Site Data":
                    continue
                    
                df_raw = pd.read_excel(excel, sheet_name=hoja, header=None)
                if df_raw.empty:
                    continue
                
                # Mapear exactamente qué filas están escondidas en el Excel real (por filtros o altura 0)
                ws_control = wb_control[hoja]
                filas_escondidas = set()
                
                # Revisar el estado de cada fila en openpyxl
                for r_idx, rd in ws_control.row_dimensions.items():
                    if rd.hidden or rd.height == 0:
                        filas_escondidas.add(r_idx - 1) # Convertir a índice base 0 de Pandas
                
                # Verificar también las filas colapsadas por el AutoFiltro dinámico de Excel
                if ws_control.auto_filter and ws_control.auto_filter.ref:
                    for row_idx in range(1, ws_control.max_row + 1):
                        row_dim = ws_control.row_dimensions.get(row_idx)
                        if row_dim and (row_dim.hidden or row_dim.height == 0):
                            filas_escondidas.add(row_idx - 1)

                fila_fechas_idx = None
                for idx, row in df_raw.iterrows():
                    # SI LA FILA ESTÁ ESCONDIDA, NO LA MIRA PARA LAS FECHAS
                    if idx in filas_escondidas:
                        continue
                        
                    valores_str = [str(val).strip() for val in row.tolist() if pd.notna(val)]
                    if any('2026' in val or '/26' in val or '-26' in val for val in valores_str):
                        fila_fechas_idx = idx
                        break
                
                if fila_fechas_idx is None:
                    continue
                
                encabezados = [str(col).strip() for col in df_raw.iloc[fila_fechas_idx].tolist()]
                
                columnas_fecha_indices = []
                fechas_objetos = []
                
                for i, enc in enumerate(encabezados):
                    if i >= 4 and not enc.startswith('Unnamed:') and enc != 'nan' and enc != '###':
                        fecha_limpia = enc.replace(' 00:00:00', '').strip()
                        
                        formatos_a_probar = ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m-%d', '%m/%d/%y', '%m/%d/%Y', '%d/%m/%y', '%d/%m/%Y']
                        dt = None
                        for fmt in formatos_a_probar:
                            try:
                                dt = datetime.strptime(fecha_limpia, fmt)
                                break
                            except:
                                continue
                        
                        if not dt and isinstance(df_raw.iloc[fila_fechas_idx, i], datetime):
                            dt = df_raw.iloc[fila_fechas_idx, i]
                            
                        if dt and dt.year == 2026:
                            columnas_fecha_indices.append(i)
                            fechas_objetos.append(dt)

                if not columnas_fecha_indices:
                    continue

                col_geo = 1       
                col_h2r = 2       
                col_programa = 3  
                col_pais = 4      
                
                for i in range(min(9, len(encabezados))):
                    enc_lower = encabezados[i].lower()
                    if 'bu' in enc_lower or 'geo' in enc_lower: col_geo = i
                    elif 'h2r' in enc_lower or 'dep.' in enc_lower: col_h2r = i
                    elif 'client' in enc_lower or 'program' in enc_lower: col_programa = i
                    elif 'country' in enc_lower or 'país' in enc_lower: col_pais = i

                df_raw[col_geo] = df_raw[col_geo].ffill()
                df_raw[col_h2r] = df_raw[col_h2r].ffill()
                df_raw[col_programa] = df_raw[col_programa].ffill()
                df_raw[col_pais] = df_raw[col_pais].ffill()
                
                registros_hoja = []
                opciones_metrica = ['Budget', 'Forecast', 'Required', 'Projected', 'Budget HC', 'Forecast HC', 'Required HC', 'Projected HC']
                
                for idx, fila in df_raw.iterrows():
                    if idx <= fila_fechas_idx:
                        continue
                        
                    # SI LA FILA ESTÁ ESCONDIDA POR FILTRO O FORMATO, SE IGNORA POR COMPLETO
                    if idx in filas_escondidas:
                        continue
                        
                    if fila.dropna().empty:
                        continue
                        
                    metrica_cruda = ""
                    for c_idx in range(min(10, len(fila))):
                        celda_val = str(fila[c_idx]).strip()
                        if celda_val in opciones_metrica:
                            metrica_cruda = celda_val
                            break
                    
                    if not metrica_cruda:
                        continue
                        
                    geo_actual = str(fila[col_geo]).strip() if pd.notna(fila[col_geo]) else "N/A"
                    h2r_actual = str(fila[col_h2r]).strip() if pd.notna(fila[col_h2r]) else "N/A"
                    programa_actual = str(fila[col_programa]).strip() if pd.notna(fila[col_programa]) else "N/A"
                    pais_actual = str(fila[col_pais]).strip() if pd.notna(fila[col_pais]) else "N/A"
                    
                    if programa_actual in ['nan', '', 'Program ID', 'Program ID (Client Name)'] or 'program id' in programa_actual.lower(): continue
                    if pais_actual in ['Country', 'nan', ''] or 'country' in pais_actual.lower(): continue
                    
                    metrica_final = metrica_cruda.replace(' HC', '').strip()
                    if metrica_final == 'Projected':
                        metrica_final = 'Project'
                        
                    for k in range(len(columnas_fecha_indices)):
                        col_idx = columnas_fecha_indices[k]
                        dt_objeto = fechas_objetos[k]
                        
                        if col_idx >= len(fila): 
                            continue
                            
                        valor_crudo = fila[col_idx]
                        valor_str = str(valor_crudo).strip()
                        
                        if valor_str == 'Closed':
                            valor_final = 'Closed'
                        elif valor_str in ['#VALUE!', 'nan', '', '111%', '105%'] or pd.isna(valor_crudo):
                            valor_final = 0
                        else:
                            try:
                                valor_num = float(valor_crudo)
                                valor_final = int(round(valor_num))
                            except:
                                valor_final = valor_str
                                
                        registros_hoja.append({
                            'Sheet_Name': hoja,
                            'Client': geo_actual if geo_actual != "nan" else hoja,  # Devuelto a Client
                            'H2R Department': h2r_actual,
                            'Program': programa_actual,
                            'Country': pais_actual,
                            'Metrica': metrica_final,
                            'Fecha_Obj': dt_objeto,
                            'Valor': valor_final
                        })
                            
                if registros_hoja:
                    dataframes_totales.append(pd.DataFrame(registros_hoja))
                    
        wb_control.close()

        if dataframes_totales:
            df_consolidado = pd.concat(dataframes_totales, ignore_index=True)
            
            df_final = df_consolidado.pivot_table(
                index=['Sheet_Name', 'Client', 'H2R Department', 'Program', 'Country', 'Fecha_Obj'],
                columns='Metrica',
                values='Valor',
                aggfunc='first'
            ).reset_index()
            
            df_final.sort_values(by=['Sheet_Name', 'Client', 'H2R Department', 'Program', 'Country', 'Fecha_Obj'], inplace=True)
            df_final['Fecha'] = df_final['Fecha_Obj'].dt.strftime('%d-%b-%y')
            
            columnas_ordenadas = ['Sheet_Name', 'Client', 'H2R Department', 'Program', 'Country', 'Fecha', 'Budget', 'Forecast', 'Required', 'Project']
            for col in columnas_ordenadas:
                if col not in df_final.columns:
                    df_final[col] = 0
                    
            df_final = df_final[columnas_ordenadas]
            
            ruta_salida = os.path.join(OUTPUT_FOLDER, 'Resumen_Miranda.xlsx')
            df_final.to_excel(ruta_salida, index=False)
            
            resultados_html = f"<h2>Excel de Miranda generado exitosamente ({len(df_final)} filas)</h2>"
            resultados_html += f"<p>El archivo se guardó en la carpeta: <b>{ruta_salida}</b></p>"
            resultados_html += df_final.head(100).to_html(index=False, border=1, na_rep="0")
            return resultados_html
        else:
            return "<h3>No se encontraron registros de métricas para el año 2026 tras aplicar los filtros.</h3>", 400
            
    except Exception as e:
        print(f"Error crítico en Miranda: {e}")
        return f"Error en el servidor: {str(e)}", 500

# =========================================================================
# 5. ENDPOINT: DEVONNE
# =========================================================================
@app.route('/upload/devonne', methods=['POST'])
def upload_devonne():
    try:
        file = request.files.get('file')
        if not file or file.filename == '':
            return "No se seleccionó ningún archivo.", 400
            
        ruta_guardado = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(ruta_guardado)
        
        print(f"--- Procesando Devonne Carl (Archivo Plano): {file.filename} ---")
        dataframes_totales = []
        
        with pd.ExcelFile(ruta_guardado, engine="openpyxl") as excel:
            for hoja in excel.sheet_names:
                sheet_state = excel.book[hoja].sheet_state
                if sheet_state == "hidden" or sheet_state == "veryHidden":
                    continue
                    
                df_raw = pd.read_excel(excel, sheet_name=hoja, header=None)
                if df_raw.empty: 
                    continue
                
                fila_fechas_idx = None
                for idx, row in df_raw.iterrows():
                    valores_str = [str(val) for val in row.tolist() if pd.notna(val)]
                    if any('2026' in val for val in valores_str):
                        fila_fechas_idx = idx
                        break
                
                if fila_fechas_idx is None: 
                    continue
                
                encabezados = [str(col).strip() for col in df_raw.iloc[fila_fechas_idx].tolist()]
                columnas_fecha_indices = []
                fechas_objetos = []
                
                for i, enc in enumerate(encabezados):
                    if i >= 3 and not enc.startswith('Unnamed:') and enc != 'nan' and enc != '###':
                        fecha_limpia = enc.replace(' 00:00:00', '').strip()
                        if '-' in fecha_limpia or '/' in fecha_limpia:
                            try:
                                dt = datetime.strptime(fecha_limpia, '%Y-%m-%d')
                                columnas_fecha_indices.append(i)
                                fechas_objetos.append(dt)
                            except:
                                try:
                                    dt = datetime.strptime(fecha_limpia, '%m-%d').replace(year=2026)
                                    columnas_fecha_indices.append(i)
                                    fechas_objetos.append(dt)
                                except:
                                    continue
                
                col_programa = None
                col_client_name = None
                col_pais = None
                col_metrica = None
                
                for i, enc in enumerate(encabezados):
                    enc_lower = enc.lower()
                    if enc_lower == 'program' or 'program\n' in enc_lower:
                        col_programa = i
                    elif 'client name' in enc_lower or 'client_name' in enc_lower or enc_lower == 'client':
                        col_client_name = i
                    elif 'country' in enc_lower or 'geo' in enc_lower: 
                        col_pais = i
                    elif 'description' in enc_lower or 'metrica' in enc_lower or 'metric' in enc_lower: 
                        col_metrica = i

                if col_programa is None: col_programa = 0
                if col_client_name is None: col_client_name = 1
                if col_pais is None: col_pais = 2
                if col_metrica is None: col_metrica = 3
                
                registros_hoja = []
                for idx, fila in df_raw.iterrows():
                    if idx <= fila_fechas_idx: 
                        continue
                        
                    metrica_cruda = str(fila[col_metrica]).strip() if pd.notna(fila[col_metrica]) else ""
                    programa_actual = str(fila[col_programa]).strip()
                    client_actual = str(fila[col_client_name]).strip()
                    pais_actual = str(fila[col_pais]).strip()
                    
                    if programa_actual in ['nan', '', 'Program', 'program id']: continue
                    if client_actual in ['nan', '', 'Client Name', 'client']: continue
                    if pais_actual in ['Country', 'Geo', 'nan', '']: continue
                    
                    if metrica_cruda in ['Budget HC', 'Forecast HC', 'Required HC', 'Projected HC', 'Budget', 'Forecast', 'Required', 'Projected']:
                        metrica_final = metrica_cruda.replace(' HC', '').strip()
                        if metrica_final == 'Projected': 
                            metrica_final = 'Project'
                            
                        for k in range(len(columnas_fecha_indices)):
                            col_idx = columnas_fecha_indices[k]
                            dt_objeto = fechas_objetos[k]
                            valor_crudo = fila[col_idx]
                            valor_str = str(valor_crudo).strip()
                            
                            if valor_str == 'Closed': 
                                valor_final = 'Closed'
                            elif valor_str in ['#VALUE!', 'nan', ''] or pd.isna(valor_crudo): 
                                valor_final = 0
                            else:
                                try: 
                                    valor_final = int(round(float(valor_crudo)))
                                except: 
                                    valor_final = valor_str
                                    
                            registros_hoja.append({
                                'Program': programa_actual,
                                'Client Name': client_actual,
                                'Country': pais_actual,
                                'Metrica': metrica_final,
                                'Fecha_Obj': dt_objeto,
                                'Valor': valor_final
                            })
                            
                if registros_hoja:
                    dataframes_totales.append(pd.DataFrame(registros_hoja))
                    
        if dataframes_totales:
            df_consolidado = pd.concat(dataframes_totales, ignore_index=True)
            
            df_final = df_consolidado.pivot_table(
                index=['Program', 'Client Name', 'Country', 'Fecha_Obj'],
                columns='Metrica',
                values='Valor',
                aggfunc='first'
            ).reset_index()
            
            df_final.sort_values(by=['Program', 'Client Name', 'Country', 'Fecha_Obj'], inplace=True)
            df_final['Fecha'] = df_final['Fecha_Obj'].dt.strftime('%d-%b-%y')
            
            df_final.rename(columns={
                'Program': 'Program H2R',
                'Client Name': 'Client'
            }, inplace=True)
            
            columnas_salida_devonne = ['Program H2R', 'Client', 'Country', 'Fecha']
            for col in ['Budget', 'Forecast', 'Required', 'Project']:
                if col in df_final.columns:
                    columnas_salida_devonne.append(col)
                    
            df_excel_devonne = df_final[columnas_salida_devonne].copy()
            nombre_base = os.path.splitext(file.filename)[0]
            ruta_excel_devonne = os.path.join(OUTPUT_FOLDER, f"{nombre_base}_Resumen_Devonne.xlsx")
            df_excel_devonne.to_excel(ruta_excel_devonne, index=False)

            html_resultado = df_excel_devonne.to_html(index=False, border=1, na_rep="")
            return f"<h2>Reporte Consolidado Devonne (Guardado en salidas_prueba)</h2><div>{html_resultado}</div>"
            
        return "<h3>No se encontraron datos para procesar en Devonne.</h3>"
        
    except Exception as e:
        return f"Error Crítico en Devonne: {str(e)}", 500


# =========================================================================
# MOTOR DE ARRANQUE ÚNICO
# =========================================================================
if __name__ == '__main__':
    print("Iniciando Servidor Unificado Completo...")
    app.run(debug=False, port=5000)